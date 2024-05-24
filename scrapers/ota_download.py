import os

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from .ota import Ota


class OtaDownload(Ota):
    """
    Class to download scraping results as excel and pdf
    """

    def __init__(self):
        super().__init__()
        self.df = None
        self.file_path = os.path.join("result", "ota.xlsx")

    def finish(self):
        self.__download_as_csv()
        self.__download_as_excel()
        self.__format_excel_file()

    def __prepare_for_download(self):
        """
        Utility: Convert to df and creating result/ directory
        """
        if self.df is None:
            self.df = pd.DataFrame(self.scraping_results)
        if not os.path.exists("result"):
            os.makedirs("result")

    def __download_as_excel(self):
        """
        Utility: Download result df as excel
        """
        self.__prepare_for_download()
        self.df.to_excel(self.file_path, sheet_name="Hotels", index=False)
        print(f"Successfully downloaded to {self.file_path}")

    def __download_as_csv(self):
        """
        Utility: Download result df as csv
        """
        self.__prepare_for_download()
        self.df.to_csv(self.file_path, sep=",", index=False)
        print(f"Successfully downloaded to {self.file_path}")

    def __format_excel_file(self):
        """
        Utility: formats the cell with different colors
        """
        wb = load_workbook(self.file_path)
        ws = wb.active

        header_columns = []

        i = 1

        while True:
            curr_cell_value = ws.cell(row=1, column=i).value
            if curr_cell_value is None:
                break
            if "Comparison" in curr_cell_value:
                header_columns.append(i)
            i += 1

        for header_column in header_columns:
            i = 2
            while True:
                cell = ws.cell(row=i, column=header_column)
                if cell.value is None:
                    break
                elif "+" in str(cell.value):
                    cell.fill = PatternFill(
                        start_color="FF9999", end_color="FF9999", fill_type="solid"
                    )
                elif "-" in str(cell.value):
                    cell.fill = PatternFill(
                        start_color="99FF99", end_color="99FF99", fill_type="solid"
                    )
                else:
                    cell.fill = PatternFill(
                        start_color="FFFF99", end_color="FFFF99", fill_type="solid"
                    )
                i += 1

        wb.save(self.file_path)
        print(f"Successfully formatted {self.file_path}")
